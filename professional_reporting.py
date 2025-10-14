"""
Professional Reporting and Documentation System for Industrial SCADA
Advanced report generation, analytics visualization, and documentation management
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union
import logging
import json
from pathlib import Path
import sqlite3
from dataclasses import dataclass
from enum import Enum
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns
from matplotlib.backends.backend_pdf import PdfPages
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.io as pio
from jinja2 import Template, Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side
from openpyxl.chart import LineChart, Reference
import io
import base64

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try to import pdfkit (optional)
try:
    import pdfkit
    PDFKIT_AVAILABLE = True
except ImportError:
    PDFKIT_AVAILABLE = False
    logger.warning("pdfkit not available - PDF generation from HTML will be disabled")

class ReportType(Enum):
    """Types of reports"""
    OPERATIONAL_SUMMARY = "operational_summary"
    PERFORMANCE_ANALYSIS = "performance_analysis"
    ALARM_SUMMARY = "alarm_summary"
    MAINTENANCE_REPORT = "maintenance_report"
    ENERGY_CONSUMPTION = "energy_consumption"
    COMPLIANCE_AUDIT = "compliance_audit"
    TREND_ANALYSIS = "trend_analysis"
    CUSTOM = "custom"

class ReportFormat(Enum):
    """Report output formats"""
    PDF = "pdf"
    EXCEL = "excel"
    HTML = "html"
    JSON = "json"
    CSV = "csv"

class ChartType(Enum):
    """Chart types for visualizations"""
    LINE_CHART = "line"
    BAR_CHART = "bar"
    PIE_CHART = "pie"
    SCATTER_PLOT = "scatter"
    HISTOGRAM = "histogram"
    HEATMAP = "heatmap"
    GAUGE = "gauge"
    TABLE = "table"

@dataclass
class ReportConfiguration:
    """Report configuration parameters"""
    report_id: str
    title: str
    report_type: ReportType
    format: ReportFormat
    time_range: Dict[str, datetime]
    data_sources: List[str]
    visualizations: List[Dict[str, Any]]
    template_path: Optional[str] = None
    auto_schedule: Optional[Dict[str, Any]] = None
    recipients: List[str] = None

@dataclass
class VisualizationConfig:
    """Visualization configuration"""
    chart_type: ChartType
    title: str
    data_query: str
    x_axis: str
    y_axis: Union[str, List[str]]
    group_by: Optional[str] = None
    aggregation: str = "mean"
    chart_options: Dict[str, Any] = None

class DataProcessor:
    """Processes raw data for reporting"""

    def __init__(self, db_connection: sqlite3.Connection):
        self.db_connection = db_connection

    def get_process_data(self, start_time: datetime, end_time: datetime,
                        point_ids: List[str] = None) -> pd.DataFrame:
        """Get process data for specified time range"""
        try:
            query = """
            SELECT point_id, timestamp, value, quality, status
            FROM monitoring_data
            WHERE timestamp BETWEEN ? AND ?
            """
            params = [start_time, end_time]

            if point_ids:
                placeholders = ','.join(['?' for _ in point_ids])
                query += f" AND point_id IN ({placeholders})"
                params.extend(point_ids)

            query += " ORDER BY timestamp"

            df = pd.read_sql_query(query, self.db_connection, params=params)
            df['timestamp'] = pd.to_datetime(df['timestamp'])

            return df

        except Exception as e:
            logger.error(f"Error getting process data: {e}")
            return pd.DataFrame()

    def get_alarm_data(self, start_time: datetime, end_time: datetime) -> pd.DataFrame:
        """Get alarm data for specified time range"""
        try:
            query = """
            SELECT alert_id, rule_id, timestamp, alert_type, priority,
                   message, source_point, current_value, acknowledged,
                   acknowledged_by, acknowledged_time, resolved, resolved_time
            FROM alerts
            WHERE timestamp BETWEEN ? AND ?
            ORDER BY timestamp DESC
            """

            df = pd.read_sql_query(query, self.db_connection, params=[start_time, end_time])
            df['timestamp'] = pd.to_datetime(df['timestamp'])

            if 'acknowledged_time' in df.columns:
                df['acknowledged_time'] = pd.to_datetime(df['acknowledged_time'])
            if 'resolved_time' in df.columns:
                df['resolved_time'] = pd.to_datetime(df['resolved_time'])

            return df

        except Exception as e:
            logger.error(f"Error getting alarm data: {e}")
            return pd.DataFrame()

    def calculate_kpis(self, process_data: pd.DataFrame) -> Dict[str, float]:
        """Calculate key performance indicators"""
        try:
            kpis = {}

            # Data availability
            total_points = len(process_data)
            good_quality = len(process_data[process_data['quality'] == 'GOOD'])
            kpis['data_availability'] = (good_quality / total_points * 100) if total_points > 0 else 0

            # Process availability (assuming system is down when no data)
            time_range = process_data['timestamp'].max() - process_data['timestamp'].min()
            expected_points = time_range.total_seconds() / 60  # Assuming 1-minute intervals
            kpis['process_availability'] = (total_points / expected_points * 100) if expected_points > 0 else 0

            # Equipment efficiency (mock calculation)
            if 'value' in process_data.columns:
                avg_performance = process_data.groupby('point_id')['value'].mean()
                kpis['average_efficiency'] = avg_performance.mean()

            # Alarm rate (alarms per hour)
            if hasattr(self, '_alarm_count'):
                hours = time_range.total_seconds() / 3600
                kpis['alarm_rate'] = (self._alarm_count / hours) if hours > 0 else 0

            return kpis

        except Exception as e:
            logger.error(f"Error calculating KPIs: {e}")
            return {}

    def aggregate_data(self, data: pd.DataFrame, time_period: str = '1H') -> pd.DataFrame:
        """Aggregate data by time period"""
        try:
            if data.empty:
                return data

            data = data.set_index('timestamp')

            numeric_columns = data.select_dtypes(include=[np.number]).columns

            aggregated = data.groupby('point_id').resample(time_period)[numeric_columns].agg({
                col: ['mean', 'min', 'max', 'std'] for col in numeric_columns
            })

            # Flatten multi-level columns
            aggregated.columns = ['_'.join(col).strip() for col in aggregated.columns]
            aggregated = aggregated.reset_index()

            return aggregated

        except Exception as e:
            logger.error(f"Error aggregating data: {e}")
            return pd.DataFrame()

class VisualizationEngine:
    """Creates professional visualizations"""

    def __init__(self):
        # Set professional style
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")

        # Configure Plotly theme
        pio.templates.default = "plotly_white"

    def create_time_series_chart(self, data: pd.DataFrame, config: VisualizationConfig) -> go.Figure:
        """Create time series line chart"""
        try:
            fig = go.Figure()

            if isinstance(config.y_axis, list):
                y_columns = config.y_axis
            else:
                y_columns = [config.y_axis]

            for y_col in y_columns:
                if y_col in data.columns:
                    fig.add_trace(go.Scatter(
                        x=data[config.x_axis],
                        y=data[y_col],
                        mode='lines+markers',
                        name=y_col,
                        line=dict(width=2),
                        marker=dict(size=4)
                    ))

            fig.update_layout(
                title=config.title,
                xaxis_title=config.x_axis,
                yaxis_title="Value",
                hovermode='x unified',
                showlegend=True,
                height=500
            )

            return fig

        except Exception as e:
            logger.error(f"Error creating time series chart: {e}")
            return go.Figure()

    def create_bar_chart(self, data: pd.DataFrame, config: VisualizationConfig) -> go.Figure:
        """Create bar chart"""
        try:
            fig = go.Figure()

            if config.group_by:
                # Grouped bar chart
                grouped_data = data.groupby(config.group_by)[config.y_axis].agg(config.aggregation).reset_index()
                fig.add_trace(go.Bar(
                    x=grouped_data[config.group_by],
                    y=grouped_data[config.y_axis],
                    name=config.y_axis
                ))
            else:
                fig.add_trace(go.Bar(
                    x=data[config.x_axis],
                    y=data[config.y_axis],
                    name=config.y_axis
                ))

            fig.update_layout(
                title=config.title,
                xaxis_title=config.x_axis,
                yaxis_title=config.y_axis,
                height=500
            )

            return fig

        except Exception as e:
            logger.error(f"Error creating bar chart: {e}")
            return go.Figure()

    def create_pie_chart(self, data: pd.DataFrame, config: VisualizationConfig) -> go.Figure:
        """Create pie chart"""
        try:
            if config.group_by:
                grouped_data = data.groupby(config.group_by)[config.y_axis].agg(config.aggregation).reset_index()
                labels = grouped_data[config.group_by]
                values = grouped_data[config.y_axis]
            else:
                labels = data[config.x_axis]
                values = data[config.y_axis]

            fig = go.Figure(data=[go.Pie(
                labels=labels,
                values=values,
                hole=0.3  # Donut chart
            )])

            fig.update_layout(
                title=config.title,
                height=500
            )

            return fig

        except Exception as e:
            logger.error(f"Error creating pie chart: {e}")
            return go.Figure()

    def create_gauge_chart(self, value: float, title: str, max_value: float = 100) -> go.Figure:
        """Create gauge chart for KPIs"""
        try:
            fig = go.Figure(go.Indicator(
                mode = "gauge+number+delta",
                value = value,
                domain = {'x': [0, 1], 'y': [0, 1]},
                title = {'text': title},
                delta = {'reference': max_value * 0.8},  # Target reference
                gauge = {
                    'axis': {'range': [None, max_value]},
                    'bar': {'color': "darkblue"},
                    'steps': [
                        {'range': [0, max_value * 0.5], 'color': "lightgray"},
                        {'range': [max_value * 0.5, max_value * 0.8], 'color': "gray"}
                    ],
                    'threshold': {
                        'line': {'color': "red", 'width': 4},
                        'thickness': 0.75,
                        'value': max_value * 0.9
                    }
                }
            ))

            fig.update_layout(height=400)
            return fig

        except Exception as e:
            logger.error(f"Error creating gauge chart: {e}")
            return go.Figure()

    def create_heatmap(self, data: pd.DataFrame, config: VisualizationConfig) -> go.Figure:
        """Create heatmap visualization"""
        try:
            # Pivot data for heatmap
            pivot_data = data.pivot_table(
                values=config.y_axis,
                index=data[config.x_axis].dt.date,
                columns=data[config.x_axis].dt.hour,
                aggfunc=config.aggregation
            )

            fig = go.Figure(data=go.Heatmap(
                z=pivot_data.values,
                x=pivot_data.columns,
                y=pivot_data.index,
                colorscale='Viridis'
            ))

            fig.update_layout(
                title=config.title,
                xaxis_title="Hour of Day",
                yaxis_title="Date",
                height=500
            )

            return fig

        except Exception as e:
            logger.error(f"Error creating heatmap: {e}")
            return go.Figure()

class ReportGenerator:
    """Generates professional reports"""

    def __init__(self, db_connection: sqlite3.Connection, output_dir: str = "reports"):
        self.db_connection = db_connection
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

        self.data_processor = DataProcessor(db_connection)
        self.viz_engine = VisualizationEngine()

        # Initialize Jinja2 template environment
        template_dir = Path("templates")
        template_dir.mkdir(exist_ok=True)
        self.jinja_env = Environment(loader=FileSystemLoader(template_dir))

    def generate_report(self, config: ReportConfiguration) -> str:
        """Generate report based on configuration"""
        try:
            logger.info(f"Generating report: {config.title}")

            # Get data for the report
            report_data = self._collect_report_data(config)

            # Generate report based on format
            if config.format == ReportFormat.PDF:
                output_path = self._generate_pdf_report(config, report_data)
            elif config.format == ReportFormat.EXCEL:
                output_path = self._generate_excel_report(config, report_data)
            elif config.format == ReportFormat.HTML:
                output_path = self._generate_html_report(config, report_data)
            elif config.format == ReportFormat.JSON:
                output_path = self._generate_json_report(config, report_data)
            else:
                raise ValueError(f"Unsupported report format: {config.format}")

            logger.info(f"Report generated successfully: {output_path}")
            return str(output_path)

        except Exception as e:
            logger.error(f"Error generating report: {e}")
            raise

    def _collect_report_data(self, config: ReportConfiguration) -> Dict[str, Any]:
        """Collect all data needed for the report"""
        try:
            start_time = config.time_range['start']
            end_time = config.time_range['end']

            # Get process data
            process_data = self.data_processor.get_process_data(
                start_time, end_time, config.data_sources
            )

            # Get alarm data
            alarm_data = self.data_processor.get_alarm_data(start_time, end_time)
            self.data_processor._alarm_count = len(alarm_data)  # For KPI calculation

            # Calculate KPIs
            kpis = self.data_processor.calculate_kpis(process_data)

            # Create visualizations
            visualizations = []
            for viz_config in config.visualizations:
                viz = self._create_visualization(process_data, viz_config)
                visualizations.append(viz)

            return {
                'config': config,
                'process_data': process_data,
                'alarm_data': alarm_data,
                'kpis': kpis,
                'visualizations': visualizations,
                'generation_time': datetime.now(),
                'data_summary': {
                    'total_data_points': len(process_data),
                    'total_alarms': len(alarm_data),
                    'time_range_hours': (end_time - start_time).total_seconds() / 3600
                }
            }

        except Exception as e:
            logger.error(f"Error collecting report data: {e}")
            return {}

    def _create_visualization(self, data: pd.DataFrame, viz_config: Dict[str, Any]) -> Dict[str, Any]:
        """Create a single visualization"""
        try:
            config = VisualizationConfig(**viz_config)

            if config.chart_type == ChartType.LINE_CHART:
                fig = self.viz_engine.create_time_series_chart(data, config)
            elif config.chart_type == ChartType.BAR_CHART:
                fig = self.viz_engine.create_bar_chart(data, config)
            elif config.chart_type == ChartType.PIE_CHART:
                fig = self.viz_engine.create_pie_chart(data, config)
            elif config.chart_type == ChartType.HEATMAP:
                fig = self.viz_engine.create_heatmap(data, config)
            else:
                logger.warning(f"Unsupported chart type: {config.chart_type}")
                return {}

            # Convert to HTML for embedding
            html_str = fig.to_html(include_plotlyjs='cdn')

            # Also save as image for PDF reports
            img_bytes = fig.to_image(format="png", width=800, height=500)
            img_b64 = base64.b64encode(img_bytes).decode()

            return {
                'config': config,
                'html': html_str,
                'image_base64': img_b64,
                'figure': fig
            }

        except Exception as e:
            logger.error(f"Error creating visualization: {e}")
            return {}

    def _generate_pdf_report(self, config: ReportConfiguration, data: Dict[str, Any]) -> Path:
        """Generate PDF report"""
        try:
            # First generate HTML report
            html_path = self._generate_html_report(config, data)

            # Convert HTML to PDF
            pdf_path = self.output_dir / f"{config.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

            options = {
                'page-size': 'A4',
                'margin-top': '0.75in',
                'margin-right': '0.75in',
                'margin-bottom': '0.75in',
                'margin-left': '0.75in',
                'encoding': "UTF-8",
                'no-outline': None,
                'enable-local-file-access': None
            }

            pdfkit.from_file(str(html_path), str(pdf_path), options=options)
            return pdf_path

        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            # Fallback: create simple PDF using matplotlib
            return self._generate_matplotlib_pdf(config, data)

    def _generate_matplotlib_pdf(self, config: ReportConfiguration, data: Dict[str, Any]) -> Path:
        """Generate PDF using matplotlib as fallback"""
        try:
            pdf_path = self.output_dir / f"{config.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

            with PdfPages(pdf_path) as pdf:
                # Title page
                fig, ax = plt.subplots(figsize=(8.5, 11))
                ax.text(0.5, 0.7, config.title, ha='center', va='center', fontsize=20, weight='bold')
                ax.text(0.5, 0.6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                       ha='center', va='center', fontsize=12)
                ax.text(0.5, 0.5, f"Time Range: {config.time_range['start']} to {config.time_range['end']}",
                       ha='center', va='center', fontsize=10)
                ax.axis('off')
                pdf.savefig(fig, bbox_inches='tight')
                plt.close()

                # KPI summary page
                kpis = data.get('kpis', {})
                if kpis:
                    fig, ax = plt.subplots(figsize=(8.5, 11))
                    ax.text(0.5, 0.9, 'Key Performance Indicators', ha='center', va='center',
                           fontsize=16, weight='bold')

                    y_pos = 0.8
                    for kpi_name, kpi_value in kpis.items():
                        ax.text(0.2, y_pos, kpi_name.replace('_', ' ').title(), ha='left', va='center', fontsize=12)
                        ax.text(0.8, y_pos, f"{kpi_value:.2f}", ha='right', va='center', fontsize=12)
                        y_pos -= 0.1

                    ax.axis('off')
                    pdf.savefig(fig, bbox_inches='tight')
                    plt.close()

            return pdf_path

        except Exception as e:
            logger.error(f"Error generating matplotlib PDF: {e}")
            raise

    def _generate_html_report(self, config: ReportConfiguration, data: Dict[str, Any]) -> Path:
        """Generate HTML report"""
        try:
            # Create HTML template if it doesn't exist
            template_path = Path("templates") / "report_template.html"
            if not template_path.exists():
                self._create_default_html_template(template_path)

            template = self.jinja_env.get_template("report_template.html")

            # Prepare template data
            template_data = {
                'title': config.title,
                'generation_time': data['generation_time'].strftime('%Y-%m-%d %H:%M:%S'),
                'time_range': {
                    'start': config.time_range['start'].strftime('%Y-%m-%d %H:%M:%S'),
                    'end': config.time_range['end'].strftime('%Y-%m-%d %H:%M:%S')
                },
                'kpis': data.get('kpis', {}),
                'data_summary': data.get('data_summary', {}),
                'visualizations': data.get('visualizations', []),
                'alarm_summary': self._create_alarm_summary(data.get('alarm_data', pd.DataFrame()))
            }

            # Render template
            html_content = template.render(**template_data)

            # Save HTML file
            html_path = self.output_dir / f"{config.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)

            return html_path

        except Exception as e:
            logger.error(f"Error generating HTML report: {e}")
            raise

    def _generate_excel_report(self, config: ReportConfiguration, data: Dict[str, Any]) -> Path:
        """Generate Excel report"""
        try:
            excel_path = self.output_dir / f"{config.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = []
                for key, value in data.get('kpis', {}).items():
                    summary_data.append({'KPI': key.replace('_', ' ').title(), 'Value': value})

                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

                # Process data sheet
                process_data = data.get('process_data', pd.DataFrame())
                if not process_data.empty:
                    process_data.to_excel(writer, sheet_name='Process Data', index=False)

                # Alarm data sheet
                alarm_data = data.get('alarm_data', pd.DataFrame())
                if not alarm_data.empty:
                    alarm_data.to_excel(writer, sheet_name='Alarms', index=False)

                # Format worksheets
                self._format_excel_worksheets(writer)

            return excel_path

        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise

    def _generate_json_report(self, config: ReportConfiguration, data: Dict[str, Any]) -> Path:
        """Generate JSON report"""
        try:
            json_path = self.output_dir / f"{config.report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

            # Prepare JSON data (exclude complex objects)
            json_data = {
                'report_info': {
                    'title': config.title,
                    'report_type': config.report_type.value,
                    'generation_time': data['generation_time'].isoformat(),
                    'time_range': {
                        'start': config.time_range['start'].isoformat(),
                        'end': config.time_range['end'].isoformat()
                    }
                },
                'kpis': data.get('kpis', {}),
                'data_summary': data.get('data_summary', {}),
                'process_data': data.get('process_data', pd.DataFrame()).to_dict('records'),
                'alarm_data': data.get('alarm_data', pd.DataFrame()).to_dict('records')
            }

            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, default=str)

            return json_path

        except Exception as e:
            logger.error(f"Error generating JSON report: {e}")
            raise

    def _create_default_html_template(self, template_path: Path):
        """Create default HTML template"""
        template_content = """
<!DOCTYPE html>
<html>
<head>
    <title>{{ title }}</title>
    <meta charset="UTF-8">
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        .header { text-align: center; margin-bottom: 30px; }
        .kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }
        .kpi-card { background: #f8f9fa; padding: 20px; border-radius: 8px; text-align: center; }
        .kpi-value { font-size: 2em; font-weight: bold; color: #007bff; }
        .section { margin: 30px 0; }
        .visualization { margin: 20px 0; text-align: center; }
        .table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        .table th, .table td { border: 1px solid #dee2e6; padding: 12px; text-align: left; }
        .table th { background: #007bff; color: white; }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ title }}</h1>
        <p>Generated: {{ generation_time }}</p>
        <p>Time Range: {{ time_range.start }} to {{ time_range.end }}</p>
    </div>

    {% if kpis %}
    <div class="section">
        <h2>Key Performance Indicators</h2>
        <div class="kpi-grid">
            {% for kpi_name, kpi_value in kpis.items() %}
            <div class="kpi-card">
                <div class="kpi-value">{{ "%.2f"|format(kpi_value) }}</div>
                <div>{{ kpi_name.replace('_', ' ').title() }}</div>
            </div>
            {% endfor %}
        </div>
    </div>
    {% endif %}

    {% if data_summary %}
    <div class="section">
        <h2>Data Summary</h2>
        <table class="table">
            {% for key, value in data_summary.items() %}
            <tr>
                <td>{{ key.replace('_', ' ').title() }}</td>
                <td>{{ value }}</td>
            </tr>
            {% endfor %}
        </table>
    </div>
    {% endif %}

    {% if visualizations %}
    <div class="section">
        <h2>Visualizations</h2>
        {% for viz in visualizations %}
        <div class="visualization">
            <h3>{{ viz.config.title }}</h3>
            {{ viz.html|safe }}
        </div>
        {% endfor %}
    </div>
    {% endif %}

    {% if alarm_summary %}
    <div class="section">
        <h2>Alarm Summary</h2>
        {{ alarm_summary|safe }}
    </div>
    {% endif %}
</body>
</html>
        """

        template_path.parent.mkdir(exist_ok=True)
        with open(template_path, 'w', encoding='utf-8') as f:
            f.write(template_content.strip())

    def _create_alarm_summary(self, alarm_data: pd.DataFrame) -> str:
        """Create alarm summary HTML"""
        if alarm_data.empty:
            return "<p>No alarms in the specified time period.</p>"

        try:
            # Alarm summary statistics
            total_alarms = len(alarm_data)
            acknowledged_alarms = len(alarm_data[alarm_data['acknowledged'] == True])
            resolved_alarms = len(alarm_data[alarm_data['resolved'] == True])

            # Priority distribution
            priority_dist = alarm_data['priority'].value_counts().to_dict()

            html = f"""
            <table class="table">
                <tr><td>Total Alarms</td><td>{total_alarms}</td></tr>
                <tr><td>Acknowledged</td><td>{acknowledged_alarms}</td></tr>
                <tr><td>Resolved</td><td>{resolved_alarms}</td></tr>
                <tr><td>Outstanding</td><td>{total_alarms - resolved_alarms}</td></tr>
            </table>

            <h3>Priority Distribution</h3>
            <table class="table">
            """

            for priority, count in priority_dist.items():
                html += f"<tr><td>{priority}</td><td>{count}</td></tr>"

            html += "</table>"

            return html

        except Exception as e:
            logger.error(f"Error creating alarm summary: {e}")
            return "<p>Error generating alarm summary.</p>"

    def _format_excel_worksheets(self, writer):
        """Format Excel worksheets"""
        try:
            workbook = writer.book

            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))

            # Format each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Format headers
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border

                # Add borders to data
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.border = border

                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        except Exception as e:
            logger.error(f"Error formatting Excel worksheets: {e}")

# Example usage and report templates
if __name__ == "__main__":
    # Initialize database connection (mock)
    import sqlite3
    db_conn = sqlite3.connect(":memory:")

    # Initialize report generator
    report_gen = ReportGenerator(db_conn)

    # Example report configuration
    config = ReportConfiguration(
        report_id="daily_operational",
        title="Daily Operational Report",
        report_type=ReportType.OPERATIONAL_SUMMARY,
        format=ReportFormat.HTML,
        time_range={
            'start': datetime.now() - timedelta(days=1),
            'end': datetime.now()
        },
        data_sources=['T001', 'P001', 'F001'],
        visualizations=[
            {
                'chart_type': ChartType.LINE_CHART,
                'title': 'Temperature Trend',
                'data_query': 'SELECT * FROM monitoring_data WHERE point_id = "T001"',
                'x_axis': 'timestamp',
                'y_axis': 'value'
            },
            {
                'chart_type': ChartType.BAR_CHART,
                'title': 'Average Values by Point',
                'data_query': 'SELECT * FROM monitoring_data',
                'x_axis': 'point_id',
                'y_axis': 'value',
                'aggregation': 'mean'
            }
        ]
    )

    try:
        # Generate report
        output_path = report_gen.generate_report(config)
        print(f"Report generated: {output_path}")

    except Exception as e:
        print(f"Error generating report: {e}")

    finally:
        db_conn.close()